#!/usr/bin/env python3
"""
WeChat receipt ingestion daemon (Windows-friendly).

Goal:
- Detect new image files continuously (no manual save click loop).
- Process with OCR.
- Extract date, time, beneficiary and amount.
- Append results to Excel with idempotency.
- Avoid missing files via periodic reconciliation scan.

Notes:
- There is no official webhook from WeChat Desktop local storage.
- This script emulates webhook behavior using filesystem events + durable queue.
"""

from __future__ import annotations

import argparse
import hashlib
import io
import json
import os
import re
import sqlite3
import sys
import threading
import time
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from PIL import Image, ImageFilter, ImageOps, ImageStat
from openpyxl import Workbook, load_workbook

try:
    from watchdog.events import FileSystemEventHandler
    from watchdog.observers import Observer

    WATCHDOG_AVAILABLE = True
except Exception:
    WATCHDOG_AVAILABLE = False
    class FileSystemEventHandler:  # type: ignore[override]
        pass
    class Observer:  # type: ignore[override]
        pass


IMG_HEADERS: dict[str, tuple[int, int]] = {
    "jpg": (0xFF, 0xD8),
    "png": (0x89, 0x50),
    "gif": (0x47, 0x49),
    "webp": (0x52, 0x49),
}

IMG_SUFFIXES = {".jpg", ".jpeg", ".png", ".bmp", ".webp", ".gif", ".dat"}
PLAIN_IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".bmp", ".webp", ".gif"}
LANCZOS_FILTER = getattr(getattr(Image, "Resampling", Image), "LANCZOS")


def is_candidate(path: Path) -> bool:
    if not path.is_file():
        return False
    if path.suffix.lower() not in IMG_SUFFIXES:
        return False

    s = str(path).lower().replace("/", "\\")

    if "\\msgattach\\" in s and "\\image\\" in s and path.suffix.lower() == ".dat":
        return True

    # WeChat can store full images in plain formats (.png/.jpg) under MsgAttach/Image.
    if "\\msgattach\\" in s and "\\image\\" in s and path.suffix.lower() in PLAIN_IMAGE_SUFFIXES:
        return True

    # Fallback lane to avoid losing incoming files when only thumbnail is available.
    if "\\msgattach\\" in s and "\\thumb\\" in s and path.suffix.lower() == ".dat":
        return True

    # Temp files do not carry stable group identity, so they are ignored
    # for CLIENTE-oriented accounting output.
    if "\\filestorage\\temp\\" in s and path.suffix.lower() in {".jpg", ".jpeg", ".png", ".bmp", ".webp"}:
        return False

    return False


def detect_source_kind(path: Path) -> str:
    s = str(path).lower().replace("/", "\\")
    if "\\msgattach\\" in s and "\\image\\" in s and path.suffix.lower() == ".dat":
        return "msgattach_image_dat"
    if "\\msgattach\\" in s and "\\thumb\\" in s and path.suffix.lower() == ".dat":
        return "msgattach_thumb_dat"
    if "\\filestorage\\temp\\" in s:
        return "temp_image"
    if "\\msgattach\\" in s and "\\image\\" in s:
        return "msgattach_image_plain"
    return "other"


def resolve_full_image_from_thumb_path(thumb_path: Path) -> Optional[Path]:
    """Try to map MsgAttach/Thumb/<month>/<hash>_t.dat -> MsgAttach/Image/<month>/<hash>.(dat|jpg|png...)."""
    s = str(thumb_path).replace("/", "\\")
    if "\\msgattach\\" not in s.lower() or "\\thumb\\" not in s.lower():
        return None

    img_loc = s.replace("\\Thumb\\", "\\Image\\").replace("\\thumb\\", "\\Image\\")
    img_path = Path(img_loc)

    stem = img_path.stem
    base = stem[:-2] if stem.lower().endswith("_t") else stem
    candidates: list[Path] = []
    for ext in (".dat", ".jpg", ".jpeg", ".png", ".webp", ".bmp", ".gif"):
        candidates.append(img_path.with_name(f"{base}{ext}"))
    for ext in (".dat", ".jpg", ".jpeg", ".png", ".webp", ".bmp", ".gif"):
        candidates.append(img_path.with_name(f"{stem}{ext}"))

    for c in candidates:
        if c.exists() and c.is_file():
            return c
    return None


def sha256_bytes(data: bytes) -> str:
    h = hashlib.sha256()
    h.update(data)
    return h.hexdigest()


def decode_wechat_dat(raw: bytes) -> tuple[bytes, str, int]:
    if len(raw) < 2:
        raise ValueError("Empty or invalid .dat file")

    for ext, (h0, h1) in IMG_HEADERS.items():
        k0 = raw[0] ^ h0
        k1 = raw[1] ^ h1
        if k0 != k1:
            continue
        key = k0
        decoded = bytes(b ^ key for b in raw)
        try:
            with Image.open(io.BytesIO(decoded)) as im:
                im.verify()
            return decoded, ext, key
        except Exception:
            continue

    raise ValueError("Unable to decode .dat as known image format")


def open_image_from_file(path: Path) -> tuple[Image.Image, bytes, str, Optional[int]]:
    raw = path.read_bytes()
    if path.suffix.lower() == ".dat":
        decoded, ext, key = decode_wechat_dat(raw)
        with Image.open(io.BytesIO(decoded)) as im:
            return im.convert("RGB"), decoded, ext, key
    with Image.open(io.BytesIO(raw)) as im:
        return im.convert("RGB"), raw, path.suffix.lower().lstrip("."), None


def quality_score(img: Image.Image) -> float:
    w, h = img.size
    gray = img.convert("L")
    var = float(ImageStat.Stat(gray).var[0])
    long_side = max(w, h)
    short_side = min(w, h)

    res_component = min(1.0, long_side / 1400.0) * 0.65 + min(1.0, short_side / 700.0) * 0.20
    contrast_component = min(1.0, var / 1800.0) * 0.15
    score = res_component + contrast_component
    return round(max(0.0, min(1.0, score)), 4)


class OCREngine:
    name = "none"

    def extract(self, img: Image.Image) -> tuple[str, float]:
        raise NotImplementedError


class RapidOCREngine(OCREngine):
    name = "rapidocr"

    def __init__(self) -> None:
        from rapidocr_onnxruntime import RapidOCR  # type: ignore

        self._ocr = RapidOCR()

    def extract(self, img: Image.Image) -> tuple[str, float]:
        import numpy as np  # type: ignore

        arr = np.array(img.convert("RGB"))
        result, _ = self._ocr(arr)
        if not result:
            return "", 0.0
        texts: list[str] = []
        confs: list[float] = []
        for item in result:
            if len(item) >= 3:
                texts.append(str(item[1]))
                try:
                    confs.append(float(item[2]))
                except Exception:
                    pass
        text = "\n".join(t for t in texts if t.strip())
        conf = (sum(confs) / len(confs)) if confs else 0.5
        return text, round(max(0.0, min(1.0, conf)), 4)


class TesseractOCREngine(OCREngine):
    name = "tesseract"

    def __init__(self) -> None:
        import pytesseract  # type: ignore

        cmd = os.getenv("TESSERACT_CMD", "").strip()
        if cmd:
            pytesseract.pytesseract.tesseract_cmd = cmd
        self._pytesseract = pytesseract
        self._lang = os.getenv("OCR_LANG", "por+eng+chi_sim")

    def extract(self, img: Image.Image) -> tuple[str, float]:
        text = self._pytesseract.image_to_string(img, lang=self._lang)
        text = text.strip()
        if not text:
            return "", 0.0
        return text, 0.55


def build_ocr_engine() -> OCREngine:
    try:
        return RapidOCREngine()
    except Exception:
        pass
    try:
        return TesseractOCREngine()
    except Exception:
        pass
    raise RuntimeError(
        "No OCR engine available. Install one:\n"
        "- pip install rapidocr-onnxruntime\n"
        "or\n"
        "- pip install pytesseract and install Tesseract OCR binary"
    )


DATE_PATTERNS = [
    re.compile(r"\b(\d{2}/\d{2}/\d{4})\b"),
    re.compile(r"\b(\d{4}-\d{2}-\d{2})\b"),
    re.compile(r"\b(\d{2}-\d{2}-\d{4})\b"),
    re.compile(r"\b(\d{2}/\d{2}/\d{2})\b"),
]
TIME_PATTERN = re.compile(r"\b(\d{2}:\d{2}(?::\d{2})?)\b")
AMOUNT_CURRENCY_PATTERN = re.compile(
    r"(R\$?|US\$|USD|BRL|CNY|RMB|¥|￥)\s*([0-9][0-9\.,]{0,20})",
    re.IGNORECASE,
)
AMOUNT_FALLBACK_PATTERN = re.compile(r"(?<!\d)([0-9]{1,3}(?:[\.,][0-9]{3})*[\.,][0-9]{2})(?!\d)")

BENEFICIARY_KEYS = [
    "favorecido",
    "beneficiario",
    "beneficiario",
    "destinatario",
    "destinatario",
    "recebedor",
    "recebedora",
    "nome",
    "recebido por",
    "para:",
    "收款方",
    "收款人",
    "对方",
]

BANK_ALLOWED = ("AMD", "DIAMOND", "CLEEND")


def normalize_text_for_match(value: str) -> str:
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.upper()
    value = re.sub(r"[^A-Z0-9]+", "", value)
    return value


def detect_bank(text: str, beneficiary: Optional[str]) -> Optional[str]:
    material = f"{text}\n{beneficiary or ''}"
    compact = normalize_text_for_match(material)
    if "DIAMOND" in compact:
        return "DIAMOND"
    if "CLEEND" in compact or "CLEND" in compact:
        return "CLEEND"
    if "AMD" in compact:
        return "AMD"
    return None


def normalize_date_for_excel(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    v = value.strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            dt = datetime.strptime(v, fmt)
            return dt.strftime("%d/%m/%Y")
        except Exception:
            continue
    return value


def normalize_time_for_excel(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    parts = value.strip().split(":")
    if len(parts) < 2:
        return None
    try:
        h = int(parts[0])
        m = int(parts[1])
    except Exception:
        return None
    if h < 0 or h > 23 or m < 0 or m > 59:
        return None
    return f"{h:02d}:{m:02d}"


def _count_date_matches(text: str) -> int:
    total = 0
    for pat in DATE_PATTERNS:
        total += len(pat.findall(text))
    return total


def looks_like_single_receipt(text: str) -> tuple[bool, str]:
    low = text.lower()
    date_count = _count_date_matches(text)
    time_count = len(TIME_PATTERN.findall(text))
    amount_count = len(AMOUNT_FALLBACK_PATTERN.findall(text)) + len(AMOUNT_CURRENCY_PATTERN.findall(text))

    has_strong_kw = any(
        kw in low
        for kw in (
            "comprovante",
            "pix",
            "transferência",
            "transferencia",
            "pagamento",
            "recibo",
            "receipt",
            "收款",
            "转账",
            "付款",
            "交易",
        )
    )

    has_table_header = (
        ("data" in low and "hora" in low and "banco" in low and "transfer" in low)
        or ("horario" in low and "banco" in low and "transfer" in low)
    )

    if has_table_header and (date_count >= 3 or time_count >= 3):
        return (False, "TABULAR_TRANSFER_LIST")
    if date_count >= 4 and amount_count >= 6:
        return (False, "MULTI_TRANSACTION_LIST")
    if not has_strong_kw and date_count >= 2 and amount_count >= 4:
        return (False, "WEAK_RECEIPT_SIGNAL")

    return (True, "OK")


def normalize_amount(value: str) -> Optional[float]:
    s = re.sub(r"[^\d,\.]", "", value.strip())
    if not s:
        return None
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        if re.search(r",\d{1,2}$", s):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    try:
        return round(float(s), 2)
    except Exception:
        return None


def prepare_image_for_ocr(img: Image.Image, source_kind: str) -> Image.Image:
    out = img.convert("RGB")
    w, h = out.size
    is_thumb_like = source_kind == "msgattach_thumb_dat" or max(w, h) <= 420
    if not is_thumb_like:
        return out

    # Miniatures are tiny and blurry; upscale + contrast helps OCR.
    if max(w, h) <= 260:
        scale = 4
    elif max(w, h) <= 420:
        scale = 3
    else:
        scale = 2

    out = out.resize((w * scale, h * scale), LANCZOS_FILTER)
    gray = out.convert("L")
    gray = ImageOps.autocontrast(gray, cutoff=2)
    gray = gray.filter(ImageFilter.MedianFilter(size=3))
    gray = gray.filter(ImageFilter.SHARPEN)
    return gray.convert("RGB")


def parse_receipt_fields(text: str, ocr_conf: float, q_score: float) -> dict[str, Any]:
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    raw = "\n".join(lines)

    txn_date: Optional[str] = None
    for pat in DATE_PATTERNS:
        m = pat.search(raw)
        if m:
            txn_date = m.group(1)
            break
    txn_date = normalize_date_for_excel(txn_date)

    txn_time: Optional[str] = None
    mt = TIME_PATTERN.search(raw)
    if mt:
        txn_time = mt.group(1)
    txn_time = normalize_time_for_excel(txn_time)

    currency: Optional[str] = None
    amount: Optional[float] = None
    amount_candidates: list[float] = []
    for m in AMOUNT_CURRENCY_PATTERN.finditer(raw):
        cur = m.group(1)
        val = normalize_amount(m.group(2))
        if val is not None:
            amount_candidates.append(val)
            currency = cur.upper().replace("US$", "USD")
    if not amount_candidates:
        for m in AMOUNT_FALLBACK_PATTERN.finditer(raw):
            val = normalize_amount(m.group(1))
            if val is not None:
                amount_candidates.append(val)
    if amount_candidates:
        amount = max(amount_candidates)
        if currency is None:
            currency = "BRL"

    beneficiary: Optional[str] = None
    low_lines = [ln.lower() for ln in lines]
    for idx, low in enumerate(low_lines):
        if any(k in low for k in BENEFICIARY_KEYS):
            original = lines[idx]
            if ":" in original:
                right = original.split(":", 1)[1].strip()
                if right:
                    beneficiary = right
                    break
            if idx + 1 < len(lines):
                nxt = lines[idx + 1].strip()
                if nxt:
                    beneficiary = nxt
                    break

    bank = detect_bank(raw, beneficiary)

    has_receipt_keyword = any(
        kw in raw.lower()
        for kw in [
            "pix",
            "comprovante",
            "transfer",
            "pagamento",
            "valor",
            "favorecido",
            "destinat",
            "recibo",
            "收款",
            "转账",
            "付款",
            "金额",
        ]
    )

    parse_conf = 0.0
    parse_conf += min(0.20, max(0.0, ocr_conf) * 0.20)
    parse_conf += 0.35 if amount is not None else 0.0
    parse_conf += 0.20 if txn_date else 0.0
    parse_conf += 0.10 if txn_time else 0.0
    parse_conf += 0.15 if beneficiary else 0.0
    parse_conf += 0.10 if bank else 0.0
    parse_conf += 0.10 if has_receipt_keyword else 0.0
    parse_conf += min(0.10, q_score * 0.10)
    parse_conf = round(min(1.0, parse_conf), 4)

    return {
        "txn_date": txn_date,
        "txn_time": txn_time,
        "beneficiary": beneficiary,
        "bank": bank,
        "amount": amount,
        "currency": currency,
        "parse_conf": parse_conf,
        "has_receipt_keyword": has_receipt_keyword,
    }


@dataclass
class QueueItem:
    file_id: str
    path: str
    source_kind: str
    ext: str
    size: int
    mtime: float
    attempts: int


def extract_group_id_from_path(path: Path) -> Optional[str]:
    parts = path.parts
    for idx, part in enumerate(parts):
        if part.lower() == "msgattach" and idx + 1 < len(parts):
            return parts[idx + 1]
    return None


class ClientResolver:
    def __init__(self, map_path: Path) -> None:
        self.map_path = map_path
        self._mtime: float = -1.0
        self._map: dict[str, str] = {}
        self.reload_if_needed(force=True)

    def _normalize_keys(self, data: dict[str, Any]) -> dict[str, str]:
        out: dict[str, str] = {}
        for k, v in data.items():
            key = str(k).strip().lower()
            val = str(v).strip()
            if key and val:
                out[key] = val
        return out

    def reload_if_needed(self, force: bool = False) -> None:
        if not self.map_path.exists():
            if force or self._map:
                self._map = {}
                self._mtime = -1.0
            return
        mtime = self.map_path.stat().st_mtime
        if not force and mtime == self._mtime:
            return
        try:
            raw = self.map_path.read_text(encoding="utf-8")
            data = json.loads(raw)
            if isinstance(data, dict):
                self._map = self._normalize_keys(data)
            else:
                self._map = {}
            self._mtime = mtime
        except Exception:
            self._map = {}
            self._mtime = mtime

    def resolve(self, source_path: Path) -> Optional[str]:
        self.reload_if_needed()
        gid = extract_group_id_from_path(source_path)
        if not gid:
            return None
        key = gid.strip().lower()
        if key in self._map:
            return self._map[key]
        return None


class StateDB:
    def __init__(self, db_path: Path) -> None:
        self.db_path = db_path
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self._conn = sqlite3.connect(str(db_path), check_same_thread=False)
        self._conn.row_factory = sqlite3.Row
        self._lock = threading.Lock()
        self._init_schema()

    def _init_schema(self) -> None:
        with self._lock:
            cur = self._conn.cursor()
            cur.executescript(
                """
                PRAGMA journal_mode=WAL;
                PRAGMA synchronous=NORMAL;

                CREATE TABLE IF NOT EXISTS files (
                    file_id TEXT PRIMARY KEY,
                    path TEXT NOT NULL,
                    source_kind TEXT NOT NULL,
                    ext TEXT NOT NULL,
                    size INTEGER NOT NULL,
                    mtime REAL NOT NULL,
                    ctime REAL NOT NULL,
                    status TEXT NOT NULL,
                    attempts INTEGER NOT NULL DEFAULT 0,
                    next_attempt REAL NOT NULL DEFAULT 0,
                    first_seen REAL NOT NULL,
                    last_seen REAL NOT NULL,
                    processed_at REAL,
                    sha256 TEXT,
                    last_error TEXT
                );
                CREATE INDEX IF NOT EXISTS idx_files_status_next ON files(status, next_attempt);
                CREATE INDEX IF NOT EXISTS idx_files_path ON files(path);

                CREATE TABLE IF NOT EXISTS receipts (
                    file_id TEXT PRIMARY KEY,
                    source_path TEXT NOT NULL,
                    source_kind TEXT NOT NULL,
                    ingested_at REAL NOT NULL,
                    sha256 TEXT NOT NULL,
                    txn_date TEXT,
                    txn_time TEXT,
                    beneficiary TEXT,
                    amount REAL,
                    currency TEXT,
                    parse_conf REAL NOT NULL,
                    quality_score REAL NOT NULL,
                    ocr_engine TEXT NOT NULL,
                    ocr_conf REAL NOT NULL,
                    ocr_chars INTEGER NOT NULL,
                    review_needed INTEGER NOT NULL,
                    ocr_text TEXT,
                    parser_json TEXT,
                    excel_sheet TEXT,
                    excel_row INTEGER
                );
                """
            )
            self._conn.commit()

    @staticmethod
    def compute_file_id(path: Path, stat: os.stat_result) -> str:
        payload = f"{str(path).lower()}|{stat.st_size}|{stat.st_mtime_ns}"
        return hashlib.sha1(payload.encode("utf-8")).hexdigest()

    def upsert_candidate(self, path: Path, settle_seconds: int, source_event: str) -> Optional[str]:
        if not is_candidate(path):
            return None
        try:
            st = path.stat()
        except FileNotFoundError:
            return None
        if st.st_size <= 0:
            return None

        now = time.time()
        file_id = self.compute_file_id(path, st)
        source_kind = detect_source_kind(path)
        ext = path.suffix.lower()
        next_attempt = now + max(1, settle_seconds)

        with self._lock:
            cur = self._conn.cursor()
            existing = cur.execute(
                """
                SELECT file_id, status
                FROM files
                WHERE path=?
                ORDER BY last_seen DESC
                LIMIT 1
                """,
                (str(path),),
            ).fetchone()
            if existing is not None and existing["status"] in ("pending", "retry", "processing"):
                cur.execute(
                    """
                    UPDATE files
                    SET last_seen=?, mtime=?, size=?, next_attempt=MIN(next_attempt, ?)
                    WHERE file_id=?
                    """,
                    (
                        float(now),
                        float(st.st_mtime),
                        int(st.st_size),
                        float(next_attempt),
                        existing["file_id"],
                    ),
                )
                self._conn.commit()
                return str(existing["file_id"])

            cur.execute(
                """
                INSERT INTO files(file_id, path, source_kind, ext, size, mtime, ctime, status, attempts, next_attempt, first_seen, last_seen)
                VALUES(?, ?, ?, ?, ?, ?, ?, 'pending', 0, ?, ?, ?)
                ON CONFLICT(file_id) DO UPDATE SET
                    last_seen=excluded.last_seen,
                    mtime=excluded.mtime,
                    size=excluded.size,
                    next_attempt=CASE
                        WHEN files.status IN ('done', 'duplicate') THEN files.next_attempt
                        ELSE MIN(files.next_attempt, excluded.next_attempt)
                    END
                """,
                (
                    file_id,
                    str(path),
                    source_kind,
                    ext,
                    int(st.st_size),
                    float(st.st_mtime),
                    float(st.st_ctime),
                    float(next_attempt),
                    float(now),
                    float(now),
                ),
            )
            self._conn.commit()
        return file_id

    def claim_next(self) -> Optional[QueueItem]:
        now = time.time()
        with self._lock:
            cur = self._conn.cursor()
            cur.execute("BEGIN IMMEDIATE")
            row = cur.execute(
                """
                SELECT file_id, path, source_kind, ext, size, mtime, attempts
                FROM files
                WHERE status IN ('pending', 'retry')
                  AND next_attempt <= ?
                ORDER BY next_attempt ASC
                LIMIT 1
                """,
                (float(now),),
            ).fetchone()
            if row is None:
                self._conn.commit()
                return None
            next_attempt_count = int(row["attempts"]) + 1
            cur.execute(
                """
                UPDATE files
                SET status='processing', attempts=?, last_error=NULL
                WHERE file_id=?
                """,
                (next_attempt_count, row["file_id"]),
            )
            self._conn.commit()
            return QueueItem(
                file_id=row["file_id"],
                path=row["path"],
                source_kind=row["source_kind"],
                ext=row["ext"],
                size=int(row["size"]),
                mtime=float(row["mtime"]),
                attempts=next_attempt_count,
            )

    def mark_done(self, file_id: str, sha256: str, processed_at: float) -> None:
        with self._lock:
            self._conn.execute(
                """
                UPDATE files
                SET status='done', processed_at=?, sha256=?, last_error=NULL
                WHERE file_id=?
                """,
                (processed_at, sha256, file_id),
            )
            self._conn.commit()

    def mark_retry(
        self,
        file_id: str,
        attempts: int,
        retry_base_sec: int,
        err: str,
        max_retries: int,
        delay_override_sec: Optional[int] = None,
    ) -> None:
        if max_retries > 0 and attempts >= max_retries:
            status = "failed"
            next_attempt = 0.0
        else:
            status = "retry"
            if delay_override_sec is not None:
                backoff = max(3, int(delay_override_sec))
            else:
                backoff = min(retry_base_sec * (2 ** max(0, attempts - 1)), 3600)
            next_attempt = time.time() + float(backoff)

        with self._lock:
            self._conn.execute(
                """
                UPDATE files
                SET status=?, next_attempt=?, last_error=?
                WHERE file_id=?
                """,
                (status, next_attempt, err[:1200], file_id),
            )
            self._conn.commit()

    def mark_hold(self, file_id: str, reason: str, delay_sec: int = 120) -> None:
        next_attempt = time.time() + max(30, delay_sec)
        with self._lock:
            self._conn.execute(
                """
                UPDATE files
                SET status='retry', next_attempt=?, last_error=?
                WHERE file_id=?
                """,
                (next_attempt, reason[:1200], file_id),
            )
            self._conn.commit()

    def receipt_exists(self, file_id: str) -> bool:
        with self._lock:
            row = self._conn.execute("SELECT 1 FROM receipts WHERE file_id=? LIMIT 1", (file_id,)).fetchone()
            return row is not None

    def requeue_mapped_missing_client(
        self,
        resolver: "ClientResolver",
        max_age_hours: int = 3,
        limit: int = 1200,
    ) -> int:
        """
        Requeue files that were previously marked with MISSING_CLIENT_MAP but now
        have a valid client mapping.
        """
        threshold = time.time() - max(1, int(max_age_hours)) * 3600
        now = time.time()
        with self._lock:
            cur = self._conn.cursor()
            rows = cur.execute(
                """
                SELECT file_id, path
                FROM files
                WHERE status='done'
                  AND last_error LIKE 'MISSING_CLIENT_MAP:%'
                  AND mtime >= ?
                ORDER BY mtime DESC
                LIMIT ?
                """,
                (float(threshold), int(max(1, limit))),
            ).fetchall()

            to_requeue: list[str] = []
            for row in rows:
                p = Path(row["path"])
                if resolver.resolve(p):
                    to_requeue.append(str(row["file_id"]))

            if not to_requeue:
                return 0

            cur.executemany(
                """
                UPDATE files
                SET status='retry', attempts=0, next_attempt=?, processed_at=NULL, last_error=NULL
                WHERE file_id=?
                """,
                [(float(now), fid) for fid in to_requeue],
            )
            self._conn.commit()
            return len(to_requeue)

    def receipt_sha_exists(self, sha256: str) -> bool:
        if not sha256:
            return False
        with self._lock:
            row = self._conn.execute("SELECT 1 FROM receipts WHERE sha256=? LIMIT 1", (sha256,)).fetchone()
            return row is not None

    def insert_receipt(self, payload: dict[str, Any]) -> None:
        with self._lock:
            self._conn.execute(
                """
                INSERT OR REPLACE INTO receipts(
                    file_id, source_path, source_kind, ingested_at, sha256,
                    txn_date, txn_time, beneficiary, amount, currency,
                    parse_conf, quality_score, ocr_engine, ocr_conf, ocr_chars,
                    review_needed, ocr_text, parser_json, excel_sheet, excel_row
                )
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    payload["file_id"],
                    payload["source_path"],
                    payload["source_kind"],
                    payload["ingested_at"],
                    payload["sha256"],
                    payload.get("txn_date"),
                    payload.get("txn_time"),
                    payload.get("beneficiary"),
                    payload.get("amount"),
                    payload.get("currency"),
                    payload["parse_conf"],
                    payload["quality_score"],
                    payload["ocr_engine"],
                    payload["ocr_conf"],
                    payload["ocr_chars"],
                    1 if payload["review_needed"] else 0,
                    payload.get("ocr_text"),
                    payload.get("parser_json"),
                    payload.get("excel_sheet"),
                    payload.get("excel_row"),
                ),
            )
            self._conn.commit()

    def close(self) -> None:
        with self._lock:
            self._conn.close()


LANC_HEADERS = [
    "CLIENTE",
    "DATA",
    "HORA",
    "BANCO",
    "VALOR",
]


class ExcelSink:
    def __init__(self, excel_path: Path) -> None:
        self.excel_path = excel_path
        self.excel_path.parent.mkdir(parents=True, exist_ok=True)
        self._lock = threading.Lock()
        self._init_workbook()

    def _sheet_has_expected_header(self, wb: Any, sheet_name: str) -> bool:
        if sheet_name not in wb.sheetnames:
            return False
        ws = wb[sheet_name]
        first_row = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), tuple())]
        return first_row == LANC_HEADERS

    def _create_fresh_workbook(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Lancamentos"
        ws.append(LANC_HEADERS)
        ws2 = wb.create_sheet("Revisar")
        ws2.append(LANC_HEADERS)
        wb.save(self.excel_path)
        wb.close()

    def _init_workbook(self) -> None:
        if not self.excel_path.exists():
            self._create_fresh_workbook()
            return

        wb = load_workbook(self.excel_path)
        ok_layout = self._sheet_has_expected_header(wb, "Lancamentos") and self._sheet_has_expected_header(wb, "Revisar")
        wb.close()

        if ok_layout:
            return

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        legacy = self.excel_path.with_name(f"{self.excel_path.stem}_legacy_{ts}{self.excel_path.suffix}")
        try:
            self.excel_path.replace(legacy)
        except Exception:
            pass
        self._create_fresh_workbook()

    def append(self, row_payload: dict[str, Any], review_needed: bool) -> tuple[str, int]:
        with self._lock:
            wb = load_workbook(self.excel_path)
            sheet = "Revisar" if review_needed else "Lancamentos"
            if sheet not in wb.sheetnames:
                ws = wb.create_sheet(sheet)
                ws.append(LANC_HEADERS)
            ws = wb[sheet]
            ws.append(
                [
                    row_payload.get("client"),
                    row_payload.get("txn_date"),
                    row_payload.get("txn_time"),
                    row_payload.get("bank"),
                    row_payload.get("amount"),
                ]
            )
            row_idx = ws.max_row
            wb.save(self.excel_path)
            wb.close()
            return (sheet, row_idx)


class IngestEventHandler(FileSystemEventHandler):  # type: ignore[misc]
    def __init__(self, db: StateDB, settle_seconds: int) -> None:
        self.db = db
        self.settle_seconds = settle_seconds

    def on_created(self, event: Any) -> None:
        if event.is_directory:
            return
        self.db.upsert_candidate(Path(event.src_path), self.settle_seconds, "created")

    def on_modified(self, event: Any) -> None:
        if event.is_directory:
            return
        self.db.upsert_candidate(Path(event.src_path), self.settle_seconds, "modified")


@dataclass
class Config:
    watch_roots: list[Path]
    db_path: Path
    excel_path: Path
    client_map_path: Path
    settle_seconds: int
    reconcile_seconds: int
    idle_sleep_seconds: float
    retry_base_seconds: int
    min_confidence: float
    max_retries: int
    thumb_wait_attempts: int
    disable_watchdog: bool


def reconcile_scan(cfg: Config, db: StateDB) -> int:
    count = 0
    now = time.time()
    # Keep reconcile bounded for plain image files to avoid ingesting very old backlog.
    plain_max_age_sec = 4 * 3600
    for root in cfg.watch_roots:
        if not root.exists():
            continue
        for p in root.rglob("*"):
            if not p.is_file():
                continue
            if not is_candidate(p):
                continue
            try:
                st = p.stat()
            except FileNotFoundError:
                continue
            s = str(p).lower().replace("/", "\\")
            if "\\msgattach\\" in s and "\\image\\" in s and p.suffix.lower() in PLAIN_IMAGE_SUFFIXES:
                if (now - float(st.st_mtime)) > plain_max_age_sec:
                    continue
            if db.upsert_candidate(p, cfg.settle_seconds, "reconcile"):
                count += 1
    return count


def process_item(
    item: QueueItem,
    db: StateDB,
    sink: ExcelSink,
    ocr: OCREngine,
    resolver: ClientResolver,
    cfg: Config,
) -> None:
    path = Path(item.path)
    source_kind = item.source_kind
    using_thumb_fallback = False
    if db.receipt_exists(item.file_id):
        db.mark_done(item.file_id, sha256="", processed_at=time.time())
        return

    if item.source_kind == "temp_image":
        db.mark_done(item.file_id, sha256="", processed_at=time.time())
        print(f"[SKIP] {path.name} | source=temp_image")
        return

    try:
        if not path.exists():
            raise FileNotFoundError(f"File disappeared: {path}")

        # Thumb is frequently too low quality for reliable OCR.
        # Prefer the corresponding full image file under MsgAttach/Image.
        if source_kind == "msgattach_thumb_dat":
            full_img = resolve_full_image_from_thumb_path(path)
            if full_img is not None:
                path = full_img
                source_kind = detect_source_kind(path)
                print(f"[INFO] {Path(item.path).name} -> using_full_image={path.name}")
            else:
                if item.attempts < cfg.thumb_wait_attempts:
                    db.mark_hold(item.file_id, reason="WAITING_FULL_IMAGE_FROM_THUMB", delay_sec=30)
                    print(f"[HOLD] {path.name} | waiting_full_image_for_ocr")
                else:
                    using_thumb_fallback = True
                    print(f"[FALLBACK] {path.name} | processing_thumb_directly")

        client = resolver.resolve(path)
        if not client:
            gid = extract_group_id_from_path(path) or "SEM_GRUPO"
            db.mark_hold(item.file_id, reason=f"MISSING_CLIENT_MAP:{gid}", delay_sec=120)
            print(f"[HOLD] {path.name} | grupo_sem_mapa={gid}")
            return

        img, img_bytes, _ext, _key = open_image_from_file(path)
        digest = sha256_bytes(img_bytes)
        if db.receipt_sha_exists(digest):
            db.mark_done(item.file_id, sha256=digest, processed_at=time.time())
            print(f"[SKIP] {path.name} | duplicate_sha")
            return
        q_score = quality_score(img)

        img_for_ocr = prepare_image_for_ocr(img, source_kind)
        text, ocr_conf = ocr.extract(img_for_ocr)
        ocr_chars = len(text)
        is_receipt, receipt_reason = looks_like_single_receipt(text)
        if not is_receipt:
            db.mark_done(item.file_id, sha256=digest, processed_at=time.time())
            print(f"[SKIP] {path.name} | not_receipt={receipt_reason}")
            return

        fields = parse_receipt_fields(text, ocr_conf=ocr_conf, q_score=q_score)
        bank = fields.get("bank")
        if bank is None:
            bank = detect_bank(f"{text}\n{client}", fields.get("beneficiary"))
            fields["bank"] = bank
        missing_core = (
            fields["amount"] is None
            and fields["txn_date"] is None
            and fields["txn_time"] is None
            and bank is None
        )
        if missing_core:
            db.mark_done(item.file_id, sha256=digest, processed_at=time.time())
            print(f"[SKIP] {path.name} | unreadable_missing_core_fields")
            return

        quality_floor = 0.20 if using_thumb_fallback else 0.38
        conf_floor = max(cfg.min_confidence, 0.70) if using_thumb_fallback else cfg.min_confidence
        review_needed = (
            fields["amount"] is None
            or fields["txn_date"] is None
            or fields["txn_time"] is None
            or bank is None
            or fields["parse_conf"] < conf_floor
            or q_score < quality_floor
        )

        payload: dict[str, Any] = {
            "file_id": item.file_id,
            "source_path": str(path),
            "source_kind": source_kind,
            "ingested_at": time.time(),
            "sha256": digest,
            "txn_date": fields["txn_date"],
            "txn_time": fields["txn_time"],
            "client": client,
            "bank": bank,
            "beneficiary": fields["beneficiary"],
            "amount": fields["amount"],
            "currency": fields["currency"],
            "parse_conf": fields["parse_conf"],
            "quality_score": q_score,
            "ocr_engine": ocr.name,
            "ocr_conf": ocr_conf,
            "ocr_chars": ocr_chars,
            "review_needed": review_needed,
            "ocr_text": text[:25000] if text else "",
            "parser_json": json.dumps(fields, ensure_ascii=False),
            "error": None,
        }

        sheet, row = sink.append(payload, review_needed=review_needed)
        payload["excel_sheet"] = sheet
        payload["excel_row"] = row
        db.insert_receipt(payload)
        db.mark_done(item.file_id, sha256=digest, processed_at=time.time())

        print(
            f"[OK] {path.name} | cliente={client} | banco={bank} | valor={fields['amount']} "
            f"| data={fields['txn_date']} {fields['txn_time']} | sheet={sheet}"
        )

    except Exception as exc:
        fast_retry = 5 if isinstance(exc, PermissionError) else None
        db.mark_retry(
            file_id=item.file_id,
            attempts=item.attempts,
            retry_base_sec=cfg.retry_base_seconds,
            err=f"{type(exc).__name__}: {exc}",
            max_retries=cfg.max_retries,
            delay_override_sec=fast_retry,
        )
        print(f"[RETRY] {path.name} | attempt={item.attempts} | err={type(exc).__name__}: {exc}")


def default_watch_roots() -> list[Path]:
    home = Path(os.environ.get("USERPROFILE", str(Path.home())))
    base = home / "Documents" / "WeChat Files"
    roots: list[Path] = []
    if base.exists():
        for sub in base.iterdir():
            fs = sub / "FileStorage"
            if fs.exists():
                roots.append(fs)
    return roots


def ensure_client_map_file(map_path: Path, watch_roots: list[Path]) -> None:
    if map_path.exists():
        return
    map_path.parent.mkdir(parents=True, exist_ok=True)

    discovered: dict[str, str] = {}
    for root in watch_roots:
        msgattach = root / "MsgAttach"
        if not msgattach.exists():
            continue
        for sub in msgattach.iterdir():
            if not sub.is_dir():
                continue
            gid = sub.name.strip()
            if gid and gid.lower() not in discovered:
                discovered[gid.lower()] = ""
            if len(discovered) >= 30:
                break
        if len(discovered) >= 30:
            break

    template: dict[str, str] = {
        "COLE_AQUI_ID_DO_GRUPO": "NOME_DO_CLIENTE",
    }
    for gid in sorted(discovered.keys()):
        template[gid] = ""
    map_path.write_text(json.dumps(template, ensure_ascii=False, indent=2), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="WeChat receipt ingestion daemon")
    p.add_argument("--watch-root", action="append", default=[], help="Root directory to monitor (repeatable)")
    p.add_argument("--db-path", default=str(Path.cwd() / "wechat_receipt_state.db"))
    p.add_argument("--excel-path", default=str(Path.cwd() / "pagamentos_wechat.xlsx"))
    p.add_argument("--client-map-path", default=str(Path.cwd() / "clientes_grupos.json"))
    p.add_argument("--settle-seconds", type=int, default=5)
    p.add_argument("--reconcile-seconds", type=int, default=90)
    p.add_argument("--idle-sleep-seconds", type=float, default=1.2)
    p.add_argument("--retry-base-seconds", type=int, default=30)
    p.add_argument("--min-confidence", type=float, default=0.55)
    p.add_argument("--max-retries", type=int, default=0, help="0 means infinite retries")
    p.add_argument("--thumb-wait-attempts", type=int, default=3)
    p.add_argument("--disable-watchdog", action="store_true")
    return p.parse_args()


def build_config(args: argparse.Namespace) -> Config:
    roots = [Path(r) for r in args.watch_root] if args.watch_root else default_watch_roots()
    return Config(
        watch_roots=roots,
        db_path=Path(args.db_path),
        excel_path=Path(args.excel_path),
        client_map_path=Path(args.client_map_path),
        settle_seconds=max(1, args.settle_seconds),
        reconcile_seconds=max(20, args.reconcile_seconds),
        idle_sleep_seconds=max(0.2, args.idle_sleep_seconds),
        retry_base_seconds=max(10, args.retry_base_seconds),
        min_confidence=max(0.0, min(1.0, args.min_confidence)),
        max_retries=max(0, args.max_retries),
        thumb_wait_attempts=max(1, int(args.thumb_wait_attempts)),
        disable_watchdog=bool(args.disable_watchdog),
    )


def main() -> int:
    args = parse_args()
    cfg = build_config(args)

    # Avoid Windows cp1252 crashes when group names contain non-Latin chars.
    try:
        if hasattr(sys.stdout, "reconfigure"):
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        if hasattr(sys.stderr, "reconfigure"):
            sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    if not cfg.watch_roots:
        print("No watch roots found. Pass --watch-root explicitly.")
        return 2

    print("Watch roots:")
    for r in cfg.watch_roots:
        print(f" - {r}")
    print(f"DB: {cfg.db_path}")
    print(f"Excel: {cfg.excel_path}")
    print(f"Client map: {cfg.client_map_path}")

    ensure_client_map_file(cfg.client_map_path, cfg.watch_roots)
    resolver = ClientResolver(cfg.client_map_path)

    db = StateDB(cfg.db_path)
    sink = ExcelSink(cfg.excel_path)
    requeued = db.requeue_mapped_missing_client(resolver, max_age_hours=3, limit=1200)
    if requeued:
        print(f"[RECOVER] requeued_missing_client={requeued}")
    try:
        ocr = build_ocr_engine()
    except Exception as exc:
        print(str(exc))
        return 3
    print(f"OCR engine: {ocr.name}")

    observer: Optional[Observer] = None
    if WATCHDOG_AVAILABLE and not cfg.disable_watchdog:
        observer = Observer()
        handler = IngestEventHandler(db, cfg.settle_seconds)
        for root in cfg.watch_roots:
            if root.exists():
                observer.schedule(handler, str(root), recursive=True)
        observer.start()
        print("Watchdog: enabled")
    else:
        print("Watchdog: disabled (using reconcile polling only)")

    last_reconcile = 0.0
    try:
        while True:
            now = time.time()
            if now - last_reconcile >= cfg.reconcile_seconds:
                added = reconcile_scan(cfg, db)
                last_reconcile = now
                print(f"[SCAN] reconcile complete | queued_or_refreshed={added}")

            item = db.claim_next()
            if item is None:
                time.sleep(cfg.idle_sleep_seconds)
                continue
            process_item(item=item, db=db, sink=sink, ocr=ocr, resolver=resolver, cfg=cfg)
    except KeyboardInterrupt:
        print("Stopping daemon...")
    finally:
        if observer is not None:
            observer.stop()
            observer.join(timeout=5)
        db.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
